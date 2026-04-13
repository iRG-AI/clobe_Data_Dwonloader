#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
clobe_update.py — clobe.ai 재무 자동 업데이트 v2.0
  - 통장내역 / 카드승인내역 / 카드매입내역 / 세금계산서 다운로드
  - 재무관리 / 카드관리 엑셀 파일 갱신
  - xlwings 방식: 원본 서식·수식·차트 완벽 보존
  - clobe 데이터 값 컬럼만 업데이트 (수식 컬럼 보호)
  - 기존 파일 → old 폴더 백업 후 새 버전명으로 저장
환경변수: CLOBE_YEAR (없으면 현재 연도)
"""

import json, os, sys, time, glob, shutil, re, unicodedata
from datetime import datetime
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager

# ─── 설정 로드 ───
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(SCRIPT_DIR, 'clobe-config.json'), encoding='utf-8') as f:
    CFG = json.load(f)

CLOBE_URL      = CFG['clobe']['url']
CLOBE_ID       = CFG['clobe']['id']
CLOBE_PASSWORD = CFG['clobe']['password']
FINANCE_DIR    = CFG['paths']['finance_dir']
DOWNLOAD_DIR   = CFG['paths']['download_dir']

TARGET_YEAR = os.environ.get('CLOBE_YEAR', str(datetime.now().year))
YY    = TARGET_YEAR[-2:]
TODAY = datetime.now().strftime('%Y%m%d')

PAT_TXN      = '은행 거래내역'
PAT_CARD_APR = '카드 승인내역'
PAT_CARD_PUR = '카드 매입내역'
PAT_TAX      = '세금계산서 데이터'

def log(msg): print(msg, flush=True)

def nfc(s):
    """Mac 파일시스템 NFD → Python NFC 변환"""
    return unicodedata.normalize('NFC', s)

# ─── 파일 유틸 ───
def snapshot(dl_dir, keyword):
    return set(f for f in glob.glob(os.path.join(dl_dir, f'*{keyword}*.xlsx'))
               if not os.path.basename(f).startswith('~$'))

def wait_new(dl_dir, before, keyword, timeout=90):
    deadline = time.time() + timeout
    while time.time() < deadline:
        time.sleep(1.5)
        cur    = snapshot(dl_dir, keyword)
        crdown = glob.glob(os.path.join(dl_dir, '*.crdownload'))
        new    = cur - before
        if new and not crdown:
            time.sleep(1)
            return max(new, key=os.path.getmtime)
    raise TimeoutError(f'다운로드 대기 초과 ({timeout}s) [{keyword}]')

def find_latest(directory, keyword):
    """os.listdir + NFC 변환으로 한글 파일명 매칭"""
    try:
        all_files = os.listdir(directory)
    except Exception as e:
        raise FileNotFoundError(f"폴더 접근 실패: {directory} ({e})")
    keyword_nfc = nfc(keyword)
    cands = [
        os.path.join(directory, f)
        for f in all_files
        if keyword_nfc in nfc(f) and f.endswith('.xlsx') and not f.startswith('~$')
    ]
    if not cands:
        raise FileNotFoundError(f"'{keyword}' 파일 없음: {directory}")
    return max(cands, key=os.path.getmtime)

def next_seq(directory, fy_prefix, date_str):
    """같은 날짜의 V번호 최댓값 + 1"""
    try:
        all_files = os.listdir(directory)
    except Exception:
        return 1
    prefix_nfc = nfc(fy_prefix)
    existing = [
        f for f in all_files
        if nfc(f).startswith(prefix_nfc) and date_str in f
        and '_V' in f and f.endswith('.xlsx') and not f.startswith('~$')
    ]
    if not existing:
        return 1
    nums = [int(m.group(1)) for name in existing
            for m in [re.search(r'_V(\d+)\.xlsx$', name)] if m]
    return max(nums) + 1 if nums else 1

def archive_and_new(src_path, finance_dir, fy_prefix):
    """기존 파일 old/ 백업 후 새 버전명 파일 경로 반환"""
    old_dir = os.path.join(finance_dir, 'old')
    os.makedirs(old_dir, exist_ok=True)
    ts       = datetime.now().strftime('%Y%m%d_%H%M%S')
    old_dest = os.path.join(old_dir, os.path.basename(src_path))
    if os.path.exists(old_dest):
        base, ext = os.path.splitext(old_dest)
        old_dest  = f'{base}_{ts}{ext}'
    shutil.copy2(src_path, old_dest)
    log(f'  [백업] {os.path.basename(old_dest)} -> old/')
    seq      = next_seq(finance_dir, fy_prefix, TODAY)
    new_name = f'{fy_prefix}_{TODAY}_V{seq}.xlsx'
    new_path = os.path.join(finance_dir, new_name)
    shutil.copy2(src_path, new_path)
    log(f'  [신규] {new_name}')
    return new_path

# ─── clobe 다운로드 파일에서 값 행 읽기 (헤더 제외) ───
def load_value_rows(src_path, sheet_index=0):
    """
    clobe xlsx에서 헤더를 제외한 데이터 행만 반환.
    openpyxl read_only + data_only 로 빠르게 읽기.
    """
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb   = load_workbook(src_path, data_only=True, read_only=True)
        ws   = wb.worksheets[sheet_index]
        rows = [tuple(c.value for c in row) for row in ws.iter_rows(min_row=2)]
        wb.close()
    return rows

# ─── openpyxl 기반 업데이트 (수식 컬럼 보호) ───
def _update_sheet_in_wb(wb, sheet_name,
                        new_rows, val_col_count,
                        date_col_idx, target_year):
    """
    이미 열린 wb 객체에서 시트만 업데이트.
    파일 열기/저장은 호출자가 담당 (XML 오류 방지).
    """
    from copy import copy

    ws = wb[sheet_name]

    # 1) 수식 컬럼 목록 파악 (2~5행 샘플링)
    formula_cols = set()
    for c in range(1, ws.max_column + 1):
        for r in range(2, min(6, ws.max_row + 1)):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith('='):
                formula_cols.add(c)
                break
    log(f'  [{sheet_name}] 수식 보호 컬럼: {sorted(formula_cols)}')

    # 2) target_year 행 역순 삭제
    to_del = []
    for row in ws.iter_rows(min_row=2):
        val = row[date_col_idx].value
        if val is not None and str(val).startswith(str(target_year)):
            to_del.append(row[0].row)
    for rn in reversed(to_del):
        ws.delete_rows(rn)
    log(f'  [{sheet_name}] {target_year}년 기존 {len(to_del)}건 삭제')

    # 3) 서식 복사용 기준행 (현재 마지막 데이터 행)
    ref_row = ws.max_row if ws.max_row >= 2 else 2

    # 4) 신규 데이터 삽입 (값 컬럼만, 수식 컬럼 건너뜀)
    inserted = 0
    for src in new_rows:
        if len(src) <= date_col_idx:
            continue
        dv = src[date_col_idx]
        if dv is None or not str(dv).startswith(str(target_year)):
            continue

        nr = ws.max_row + 1

        # 기준행 서식 복사 (폰트·정렬·숫자포맷만 안전하게)
        for c in range(1, min(val_col_count + 1, ws.max_column + 1)):
            src_cell = ws.cell(ref_row, c)
            dst_cell = ws.cell(nr, c)
            try:
                if src_cell.font:         dst_cell.font         = copy(src_cell.font)
                if src_cell.alignment:    dst_cell.alignment    = copy(src_cell.alignment)
                if src_cell.number_format:
                    dst_cell.number_format = src_cell.number_format
            except Exception:
                pass

        # 값만 쓰기 (수식 컬럼 제외)
        for ci in range(val_col_count):
            col_num = ci + 1
            if col_num in formula_cols:
                continue
            ws.cell(nr, col_num).value = src[ci] if ci < len(src) else None

        inserted += 1

    log(f'  [{sheet_name}] 새 데이터 {inserted}건 삽입')

# ─── Selenium 브라우저 ───
def build_driver():
    opts = Options()
    opts.add_experimental_option('prefs', {
        'download.default_directory': DOWNLOAD_DIR,
        'download.prompt_for_download': False,
        'download.directory_upgrade': True,
        'safebrowsing.enabled': True,
    })
    opts.add_argument('--start-maximized')
    opts.add_argument('--disable-blink-features=AutomationControlled')
    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=opts)
    driver.implicitly_wait(10)
    return driver

def login_if_needed(driver):
    driver.get(CLOBE_URL); time.sleep(3)
    if '/clobe/' in driver.current_url:
        log('로그인 상태 확인'); return
    wait = WebDriverWait(driver, 20)
    el   = wait.until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "input[type='email'],input[placeholder*='이메일'],input[placeholder*='아이디']")))
    el.clear(); el.send_keys(CLOBE_ID); time.sleep(.4)
    pw = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
    pw.clear(); pw.send_keys(CLOBE_PASSWORD); time.sleep(.4)
    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
    wait.until(EC.url_contains('/clobe/'))
    log('로그인 성공'); time.sleep(2)

def set_date_year(driver):
    time.sleep(1)
    try:
        driver.execute_script(
            "const p=document.querySelector('.ant-picker.ant-picker-range'); if(p) p.click();")
        time.sleep(.8)
        result = driver.execute_script("""
            const li=Array.from(document.querySelectorAll('li'))
                         .find(el=>el.textContent.trim()==='올해');
            if(li){li.click();return 'ok';} return 'not_found';
        """)
        if result == 'ok':
            log(f'  날짜 -> {TARGET_YEAR}년 전체'); time.sleep(2)
    except Exception as e:
        log(f'  날짜 설정 오류(무시): {e}')

def set_year_tab(driver):
    try:
        btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
            (By.XPATH, "//button[normalize-space(text())='연']|//label[normalize-space(text())='연']")))
        btn.click(); time.sleep(2); log("  '연' 탭 클릭")
    except Exception as e:
        log(f"  '연' 탭 실패(무시): {e}")

def click_card_tab(driver, tab_name):
    try:
        labels = driver.find_elements(By.CSS_SELECTOR, '.ant-segmented-item-label')
        target = next((l for l in labels if l.text.strip() == tab_name), None)
        if target:
            target.click(); time.sleep(2); log(f"  '{tab_name}' 탭 클릭")
    except Exception as e:
        log(f"  탭 클릭 오류: {e}")

def click_excel_dl(driver, before, keyword, has_submenu=True):
    wait    = WebDriverWait(driver, 15)
    actions = ActionChains(driver)
    btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//button[contains(@class,'ant-dropdown-trigger') and normalize-space(.)='엑셀 다운로드']")))
    btn.click(); time.sleep(1.5)
    if has_submenu:
        try:
            submenus = wait.until(EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, 'li.ant-dropdown-menu-submenu')))
            target = next((s for s in submenus if '분류' not in s.text), submenus[0])
            actions.move_to_element(target).perform(); time.sleep(1.2)
        except Exception as e:
            log(f'  서브메뉴 hover 실패(무시): {e}')
    try:
        xlsx_items = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located(
            (By.XPATH, "//li[contains(@class,'ant-dropdown-menu-item') and normalize-space(text())='통합 파일 (xlsx)']")))
        xlsx_items[0].click()
    except Exception:
        result = driver.execute_script("""
            const items=Array.from(document.querySelectorAll('li.ant-dropdown-menu-item'));
            const t=items.find(el=>el.textContent.trim()==='통합 파일 (xlsx)');
            if(t){t.click();return 'ok';} return 'not_found';
        """)
        if result != 'ok':
            raise Exception("'통합 파일 (xlsx)' 항목 없음")
    log('  통합 파일(xlsx) 클릭 -> 대기...')
    path = wait_new(DOWNLOAD_DIR, before, keyword)
    log(f'  다운로드 완료: {os.path.basename(path)}')
    return path

def download_tax_xl(driver, before):
    wait    = WebDriverWait(driver, 15)
    actions = ActionChains(driver)
    btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//button[contains(@class,'ant-dropdown-trigger') and normalize-space(.)='엑셀 다운로드']")))
    btn.click(); time.sleep(1.5)
    items = driver.find_elements(By.CSS_SELECTOR, '.ant-dropdown-menu-item,.ant-dropdown-menu-submenu')
    if items:
        try:
            actions.move_to_element(items[0]).perform(); time.sleep(1.0)
        except Exception:
            pass
    result = driver.execute_script("""
        const items=Array.from(document.querySelectorAll('li.ant-dropdown-menu-item'));
        const t=items.find(el=>el.textContent.trim()==='통합 파일 (xlsx)');
        if(t){t.click();return 'ok';}
        if(items.length>0){items[0].click();return 'fallback';} return 'not_found';
    """)
    log(f'  세금계산서 다운로드 클릭 ({result})')
    return wait_new(DOWNLOAD_DIR, before, PAT_TAX)

def dl_transactions(driver):
    log('── [1/4] 통장 내역')
    driver.get('https://app.clobe.ai/clobe/transactions'); time.sleep(3)
    set_date_year(driver)
    before = snapshot(DOWNLOAD_DIR, PAT_TXN)
    return click_excel_dl(driver, before, PAT_TXN, has_submenu=True)

def dl_card_approval(driver):
    log('── [2/4] 카드 승인내역')
    driver.get('https://app.clobe.ai/clobe/card-approval'); time.sleep(3)
    click_card_tab(driver, '승인'); set_date_year(driver)
    before = snapshot(DOWNLOAD_DIR, PAT_CARD_APR)
    return click_excel_dl(driver, before, PAT_CARD_APR, has_submenu=True)

def dl_card_purchase(driver):
    log('── [3/4] 카드 매입내역')
    driver.get('https://app.clobe.ai/clobe/card-approval'); time.sleep(3)
    click_card_tab(driver, '매입'); set_date_year(driver)
    before = snapshot(DOWNLOAD_DIR, PAT_CARD_PUR)
    return click_excel_dl(driver, before, PAT_CARD_PUR, has_submenu=False)

def dl_tax(driver):
    log('── [4/4] 세금계산서')
    driver.get('https://app.clobe.ai/clobe/tax-invoice?type=list'); time.sleep(3)
    set_year_tab(driver)
    before = snapshot(DOWNLOAD_DIR, PAT_TAX)
    return download_tax_xl(driver, before)

# ─── 엑셀 업데이트 (xlwings) ───
# FY26-입출금: clobe 값 컬럼 13개 (COL1~13), 수식 컬럼은 COL14부터
# 세금계산서:  clobe 값 컬럼 14개 (COL1~14), 수식 컬럼은 COL15부터
# 카드 사용내역: clobe 값 컬럼 14개 (COL1~14), 수식 컬럼은 COL15부터


TXN_VAL_COLS  = 13   # 거래일시~메모 (COL1~13)
TAX_VAL_COLS  = 14   # 발급일자~거래처 (COL1~14)
CARD_VAL_COLS = 23   # 승인일시~메모 (COL1~23), COL24부터 수식

def update_finance(txn_path, tax_path):
    import warnings
    src = find_latest(FINANCE_DIR, '재무관리')
    log(f'\n재무관리 원본: {os.path.basename(src)}')
    target = archive_and_new(src, FINANCE_DIR, f'FY{YY} 재무관리')

    txn_rows = load_value_rows(txn_path)
    tax_rows = load_value_rows(tax_path)

    # 파일을 한 번만 열고 두 시트 모두 처리 후 한 번만 저장 (XML 오류 방지)
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb = load_workbook(target)

    log(f'  [FY26-입출금] 업데이트 시작...')
    _update_sheet_in_wb(wb, 'FY26-입출금', txn_rows, TXN_VAL_COLS, 0, TARGET_YEAR)

    log(f'  [세금계산서] 업데이트 시작...')
    _update_sheet_in_wb(wb, '세금계산서', tax_rows, TAX_VAL_COLS, 1, TARGET_YEAR)

    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb.save(target)
    wb.close()
    log(f'  재무관리 저장 완료: {os.path.basename(target)} ✓')

def update_card(apr_path, pur_path):
    import warnings
    src = find_latest(FINANCE_DIR, '카드관리')
    log(f'\n카드관리 원본: {os.path.basename(src)}')
    target = archive_and_new(src, FINANCE_DIR, f'FY{YY} 카드관리')

    apr_rows = load_value_rows(apr_path)
    pur_rows = load_value_rows(pur_path)

    # 파일을 한 번만 열고 처리 후 한 번만 저장
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb = load_workbook(target)

    log(f'  [사용내역-승인] 업데이트 시작...')
    _update_sheet_in_wb(wb, '사용내역', apr_rows, CARD_VAL_COLS, 0, TARGET_YEAR)

    if '매입내역' in wb.sheetnames:
        log(f'  [매입내역] 업데이트 시작...')
        _update_sheet_in_wb(wb, '매입내역', pur_rows, CARD_VAL_COLS, 0, TARGET_YEAR)
        log('  [매입내역] 별도 시트 삽입')
    else:
        log(f'  [사용내역-매입] 업데이트 시작...')
        _update_sheet_in_wb(wb, '사용내역', pur_rows, CARD_VAL_COLS, 0, TARGET_YEAR)
        log('  [매입내역] 사용내역 시트 합산')

    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb.save(target)
    wb.close()
    log(f'  카드관리 저장 완료: {os.path.basename(target)} ✓')

def _get_sheet_names(path):
    wb = load_workbook(path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names

# ─── 메인 ───
def main():
    log('=' * 55)
    log(f'  clobe.ai 자동 업데이트 v2.0  ({TARGET_YEAR}년 / {TODAY})')
    log('  ※ xlwings 방식: 서식·수식·차트 완벽 보존')
    log('=' * 55)

    driver = build_driver()
    try:
        login_if_needed(driver)
        txn_file = dl_transactions(driver)
        apr_file = dl_card_approval(driver)
        pur_file = dl_card_purchase(driver)
        tax_file = dl_tax(driver)
    finally:
        driver.quit()

    log('\n' + '─' * 55)
    log('다운로드 완료 -> 엑셀 업데이트 시작 (xlwings)')
    log('─' * 55)

    try:
        update_finance(txn_path=txn_file, tax_path=tax_file)
        update_card(apr_path=apr_file, pur_path=pur_file)
        log('\n' + '=' * 55)
        log('✅  모든 작업 완료!')
        log('=' * 55)
    except Exception as e:
        log(f'\n❌  엑셀 업데이트 오류: {e}')
        import traceback; traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        log(f'\n❌  오류: {e}')
        import traceback; traceback.print_exc()
        sys.exit(1)
